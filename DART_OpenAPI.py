"""
=============================================================================
DART 기업 정보 및 재무제표 검색 도구 (DART Corporation & Financial Statement Search Tool)
=============================================================================

이 프로그램은 한국 금융감독원의 DART 시스템에서 기업 정보와 재무제표를 검색하고 추출합니다.

주요 기능:
1. 기업명 키워드로 기업 검색
2. 기업코드로 직접 검색
3. 특정 기간의 재무보고서 목록 조회
4. 재무제표(재무상태표, 손익계산서 등) 추출 및 엑셀 저장

사용 예시:
- 기업 업데이트: python DART_OpenAPI.py --update
- 기업 검색: python DART_OpenAPI.py --keyword 삼성
- 보고서 목록: python DART_OpenAPI.py --corp-code 00126380 --bgn-de 20200101 --list-reports
- 재무제표 추출: python DART_OpenAPI.py --corp-code 00126380 --bgn-de 20200101 --extract-fs --save-excel output.xlsx

=============================================================================
"""

import argparse
import json
import os
import re
import sys
from typing import Any, Dict, List, Optional

# Global variable for dart_fss module
dart = None

try:
    import dart_fss as imported_dart
    dart = imported_dart
except ImportError:
    pass

# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# FinancialStatementExporter class - Exporter Pattern
# ---------------------------------------------------------------------------
class FinancialStatementExporter:
    """
    추출된 FinancialStatement 객체를 커스텀 서식과 스타일이 입혀진
    엑셀 파일 등으로 정교하게 변환하여 저장하는 내보내기 도구 클래스입니다.
    """
    def __init__(self, financial_statement):
        self.statement = financial_statement

    def save(self, file_path: str, include_labels: bool = False, strip_prefix: bool = True) -> str:
        """
        재무제표 데이터를 커스텀 스타일이 입혀진 엑셀 파일로 저장합니다.
        """
        import os
        import pandas as pd
        import openpyxl.styles
        from openpyxl.utils import get_column_letter

        # file_path의 폴더가 존재하지 않는 경우 생성
        dir_name = os.path.dirname(file_path)
        if dir_name:
            os.makedirs(dir_name, exist_ok=True)

        with pd.ExcelWriter(file_path) as writer:
            # 1. 회사 기본 정보 저장
            info_df = pd.DataFrame({"info": self.statement.info})
            info_df.to_excel(writer, sheet_name="info")
            
            # 2. 재무제표 종류별로 시트 저장
            for tp in self.statement._statements:
                fs_df = self.statement._statements[tp]
                if fs_df is not None:
                    sheet_name = tp if strip_prefix else "Data_" + tp
                    fs_df.to_excel(writer, sheet_name=sheet_name)
                    
                    # 3. 년도별 수치 컬럼에 커스텀 스타일 및 쉼표 서식 적용
                    ws = writer.sheets[sheet_name]
                    num_idx_levels = fs_df.index.nlevels
                    for col_idx, col_name in enumerate(fs_df.columns):
                        col_level_0 = col_name[0] if isinstance(col_name, tuple) else col_name
                        if isinstance(col_level_0, str) and any(c.isdigit() for c in col_level_0) and all(c.isdigit() or c in '-./ ' for c in col_level_0):
                            excel_col_idx = col_idx + 1 + num_idx_levels
                            col_letter = get_column_letter(excel_col_idx)
                            
                            # 열 너비 고정(20)
                            ws.column_dimensions[col_letter].width = 20
                            
                            # 회계 쉼표 서식 적용
                            start_row = fs_df.columns.nlevels + 1
                            for row in range(start_row, ws.max_row + 1):
                                cell = ws.cell(row=row, column=excel_col_idx)
                                cell.number_format = r'_-* #,##0_-;-* #,##0_-;_-* "-"_-;_-@_-'
                    
                    # 4. 계정명 매핑용 Labels 시트 선택적 저장
                    if include_labels:
                        label_sheet_name = "Labels_" + tp
                        label_df = self.statement._labels[tp]
                        label_df.to_excel(writer, sheet_name=label_sheet_name)
                        
        return file_path


def initialize_environment(force_check: bool = False):
    """
    애플리케이션 구동에 필요한 패키지 정합성을 검증하고 환경을 초기화합니다.
    """
    global dart
    if dart is not None and not force_check:
        return

    # 1. dart-fss 패키지 확인 및 동적 설치
    try:
        import dart_fss as imported_dart
        dart = imported_dart
    except ImportError:
        print('dart-fss package not found. Attempting to install dart-fss...', file=sys.stderr)
        try:
            import subprocess
            subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'dart-fss>=0.4.16'])
            import importlib
            dart = importlib.import_module('dart_fss')
            print('dart-fss installed successfully.', file=sys.stderr)
        except Exception as _err:
            print('Automatic installation failed. Please install manually and re-run:', file=sys.stderr)
            print('  pip install -r requirements.txt', file=sys.stderr)
            print('or', file=sys.stderr)
            print('  pip install dart-fss', file=sys.stderr)
            print(f'Error detail: {_err}', file=sys.stderr)
            sys.exit(1)

    # 2. openpyxl 패키지 확인 및 동적 설치
    try:
        import openpyxl  # type: ignore
    except ImportError:
        try:
            print('openpyxl not found. Attempting to install openpyxl...', file=sys.stderr)
            import subprocess
            subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl'])
            import importlib
            importlib.import_module('openpyxl')
            print('openpyxl installed successfully.', file=sys.stderr)
        except Exception as _err:
            print('Could not install openpyxl automatically. --save-excel may fail.', file=sys.stderr)
            print('Install manually with: pip install openpyxl', file=sys.stderr)


# 기본 Open DART API 키 (환경변수 DART_API_KEY 또는 --apikey로 재정의 가능)
DEFAULT_API_KEY = 'a494746598d98ab710d7d4a4da14a8936497bd8c'
CACHE_FILE_NAME = 'corp_list_cache.json'


def set_api_key(api_key: Optional[str] = None):
    global dart
    if dart is None:
        initialize_environment()
    api_key = api_key or os.environ.get('DART_API_KEY') or DEFAULT_API_KEY
    if not api_key:
        raise ValueError(
            'Open DART API key is required. Set DART_API_KEY environment variable or use --apikey.'
        )
    dart.set_api_key(api_key=api_key)


def build_cache_path(cache_file: Optional[str] = None) -> str:
    if cache_file:
        return os.path.abspath(cache_file)
    base_dir = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_dir, CACHE_FILE_NAME)


def extract_corp_info(corp) -> Dict[str, Optional[str]]:
    return {
        'corp_code': corp.corp_code,
        'corp_name': corp.corp_name,
        'stock_code': corp.stock_code,
        'corp_eng_name': getattr(corp, 'corp_eng_name', None),
        'corp_cls': getattr(corp, 'corp_cls', None),
        'modify_date': corp.modify_date,
    }


def download_corp_list(api_key: Optional[str] = None, cache_file: Optional[str] = None, force: bool = False) -> List[Dict[str, Optional[str]]]:
    """Download the current DART corporation code list and cache it locally."""
    set_api_key(api_key)
    cache_path = build_cache_path(cache_file)

    if os.path.exists(cache_path) and not force:
        return load_cached_corp_list(cache_file=cache_path)

    corp_list = dart.get_corp_list()
    corps = getattr(corp_list, 'corps', None) or []
    corp_info = [extract_corp_info(corp) for corp in corps]

    with open(cache_path, 'w', encoding='utf-8') as fp:
        json.dump(corp_info, fp, ensure_ascii=False, indent=2)

    return corp_info


def load_cached_corp_list(cache_file: Optional[str] = None) -> List[Dict[str, Optional[str]]]:
    cache_path = build_cache_path(cache_file)
    if not os.path.exists(cache_path):
        return []
    with open(cache_path, 'r', encoding='utf-8') as fp:
        return json.load(fp)


def search_corp_keyword(
    keyword: str,
    api_key: Optional[str] = None,
    cache_file: Optional[str] = None,
    market: Optional[str] = None,
    max_results: int = 20,
    force_update: bool = False,
) -> List[Dict[str, Optional[str]]]:
    """Search the cached corp list by keyword and return matching suggestions."""
    if force_update:
        corp_list = download_corp_list(api_key=api_key, cache_file=cache_file, force=True)
    else:
        corp_list = load_cached_corp_list(cache_file=cache_file)
        if not corp_list:
            corp_list = download_corp_list(api_key=api_key, cache_file=cache_file, force=False)

    regex = re.compile(re.escape(keyword), re.IGNORECASE)
    market_filter = {m.upper() for m in market} if market else None

    results = []
    for corp in corp_list:
        corp_cls = corp.get('corp_cls')
        if market_filter and corp_cls is not None and corp_cls not in market_filter:
            continue
        if regex.search(corp.get('corp_name', '') or '') or regex.search(corp.get('corp_code', '') or ''):
            results.append(corp)
            if len(results) >= max_results:
                break

    return results


def get_corp_object(corp_code: str, api_key: Optional[str] = None):
    """Retrieve a Corp object by corp_code from DART."""
    set_api_key(api_key)
    corp_list = dart.get_corp_list()
    corp = corp_list.find_by_corp_code(corp_code)
    if corp is None:
        raise ValueError(f'Corporation with code {corp_code} not found')
    return corp


def get_available_financial_reports(
    corp_code: str,
    start_date: str,
    end_date: Optional[str] = None,
    api_key: Optional[str] = None,
    report_type: Optional[str] = None,
    page_count: int = 100,
    disclosure_type: Optional[str] = None,
) -> List[Dict[str, str]]:
    """
    기업의 재무보고서 목록을 조회합니다.
    (Search for available financial reports for a corporation within a date range)
    
    필수 인자:
    -----------
    corp_code : str
        DART 기업코드 (8자리)
    start_date : str
        검색 시작일자 (YYYYMMDD 형식)
    
    선택 인자:
    -----------
    end_date : str, 선택사항
        검색 종료일자 (YYYYMMDD 형식). 지정하지 않으면 오늘 기준
    api_key : str, 선택사항
        Open DART API 키. 미지정시 기본값 사용
    report_type : str, 선택사항
        보고서 유형 필터. 예) 'a001' (사업보고서), 'a002' (반기), 'a003' (분기)
    page_count : int, 선택사항
        최대 검색 건수 (기본값: 100)
    disclosure_type : str, 선택사항
        공시유형 필터. 예) 'A' (정기공시), 'B' (주요사항보고) 등
        
    반환값:
    --------
    list of dict
        각 보고서: rcept_no, report_nm, corp_name, reprt_code, bgn_de, end_de
    """
    corp = get_corp_object(corp_code, api_key=api_key)
    
    search_kwargs = {'bgn_de': start_date, 'page_count': page_count}
    if end_date:
        search_kwargs['end_de'] = end_date
    if report_type:
        search_kwargs['pblntf_detail_ty'] = report_type
    if disclosure_type:
        search_kwargs['pblntf_ty'] = disclosure_type
    
    reports = corp.search_filings(**search_kwargs)
    
    if reports is None:
        return []
    
    result = []
    for report in reports.report_list:
        reprt_code = getattr(report, 'reprt_code', None)
        if reprt_code is None:
            reprt_code = getattr(report, 'report_code', None)
        if reprt_code is None:
            reprt_code = getattr(report, 'report_tp', None)

        result.append({
            'rcept_no': getattr(report, 'rcept_no', None),
            'report_nm': getattr(report, 'report_nm', None),
            'corp_name': getattr(report, 'corp_name', None),
            'reprt_code': reprt_code,
            'bgn_de': getattr(report, 'bgn_de', None),
            'end_de': getattr(report, 'end_de', None),
        })
    
    return result


def extract_financial_statement(
    corp_code: str,
    start_date: str,
    end_date: Optional[str] = None,
    financial_statement_types: tuple = ('bs', 'is', 'cis', 'cf'),
    separate: bool = False,
    report_type: str = 'annual',
    lang: str = 'ko',
    separator: bool = True,
    dataset: str = 'xbrl',
    cumulative: bool = False,
    api_key: Optional[str] = None,
) -> Any:
    """
    기업의 재무제표를 추출합니다.
    (Extract financial statement data for a corporation)
    
    필수 인자:
    -----------
    corp_code : str
        DART 기업코드 (8자리)
    start_date : str
        검색 시작일자 (YYYYMMDD 형식)
    
    선택 인자:
    -----------
    end_date : str, 선택사항
        검색 종료일자 (YYYYMMDD 형식). 지정하지 않으면 오늘 기준
    financial_statement_types : tuple of str, 선택사항
        추출할 재무제표 유형 (쉼표로 구분):
        'bs'  = 재무상태표 (Balance Sheet)
        'is'  = 손익계산서 (Income Statement)
        'cis' = 포괄손익계산서 (Comprehensive Income Statement)
        'cf'  = 현금흐름표 (Cash Flow Statement)
        기본값: ('bs', 'is', 'cis', 'cf')
    separate : bool, 선택사항
        True면 개별재무제표, False면 연결재무제표 (기본값: False)
    report_type : str, 선택사항
        보고서 유형: 'annual' (연간), 'half' (반기), 'quarter' (분기)
    lang : str, 선택사항
        언어: 'ko' (한글, 기본값), 'en' (영문)
    separator : bool, 선택사항
        1000단위 구분자 표시 여부 (기본값: True)
    dataset : str, 선택사항
        데이터 우선순위: 'xbrl' (XBRL 파일 우선, 기본값), 'web' (웹페이지 우선)
    api_key : str, 선택사항
        Open DART API 키. 미지정시 기본값 사용
        
    반환값:
    --------
    dict
        추출된 재무제표 객체 (pandas DataFrame으로 변환 또는 엑셀로 저장 가능)
    """
    corp = get_corp_object(corp_code, api_key=api_key)
    
    fs = corp.extract_fs(
        bgn_de=start_date,
        end_de=end_date,
        fs_tp=financial_statement_types,
        separate=separate,
        report_tp=report_type,
        lang=lang,
        separator=separator,
        dataset=dataset,
        cumulative=cumulative,
        progressbar=True,
        skip_error=True,
    )
    
    return fs


def download_original_document(
    rcept_no: str,
    save_path: str,
    api_key: Optional[str] = None,
) -> str:
    """
    공시정보의 원본 보고서(ZIP 파일)를 다운로드합니다.
    
    필수 인자:
    -----------
    rcept_no : str
        공시 접수번호 (14자리)
    save_path : str
        저장할 파일 경로 (폴더 지정 시 [접수번호].zip 으로 자동 저장)
    api_key : str, 선택사항
        Open DART API 키. 미지정시 기본값 사용
        
    반환값:
    --------
    str
        저장된 파일의 절대 경로
    """
    api_key = api_key or os.environ.get('DART_API_KEY') or DEFAULT_API_KEY
    if not api_key:
        raise ValueError(
            'Open DART API key is required. Set DART_API_KEY environment variable or use --apikey.'
        )
    
    import requests
    
    url = "https://opendart.fss.or.kr/api/document.xml"
    params = {
        'crtfc_key': api_key,
        'rcept_no': rcept_no
    }
    
    response = requests.get(url, params=params, stream=True)
    if response.status_code != 200:
        raise Exception(f"HTTP Error: {response.status_code}")
        
    content = response.content
    if not content.startswith(b'PK\x03\x04'):
        # API 에러 확인 (XML 형식)
        try:
            import xml.etree.ElementTree as ET
            root = ET.fromstring(content)
            status = root.findtext('status')
            message = root.findtext('message')
            if status and message:
                raise Exception(f"DART API Error {status}: {message}")
        except ET.ParseError:
            pass
        raise Exception("다운로드에 실패했습니다: 유효한 ZIP 파일이 아니거나 API 에러가 발생했습니다.")
        
    # 저장경로 처리
    is_dir = os.path.isdir(save_path) or save_path.endswith('/') or save_path.endswith('\\')
    save_path = os.path.abspath(save_path)
    if is_dir:
        os.makedirs(save_path, exist_ok=True)
        save_path = os.path.join(save_path, f"{rcept_no}.zip")
    else:
        dir_name = os.path.dirname(save_path)
        if dir_name:
            os.makedirs(dir_name, exist_ok=True)
        
    with open(save_path, 'wb') as f:
        f.write(content)
        
    return save_path


def print_corp_summary(corp: Dict[str, Optional[str]]):
    print('---')
    for key, value in corp.items():
        print(f'{key}: {value}')


class DocumentBatchService:
    """
    DART 공시보고서 원본 다운로드 및 재무제표 일괄 추출 등의
    복합적인 비즈니스 로직을 제공하는 서비스 클래스입니다.
    """
    def __init__(self, api_key: Optional[str] = None):
        self.api_key = api_key

    def download_and_compress_reports(
        self,
        selected_items: List[Dict[str, Any]],
        base_dir: str,
        progress_callback=None
    ) -> str:
        """
        선택된 여러 공시 보고서를 개별 다운로드한 뒤 하나의 ZIP 파일로 결합하여 저장하고,
        생성된 최종 ZIP 파일의 경로를 반환합니다.
        
        progress_callback: (current_idx, total_count, message) 형식으로 진행률을 전달하는 콜백
        """
        import shutil
        import zipfile

        temp_dir = os.path.join(base_dir, "temp")
        if os.path.exists(temp_dir):
            try:
                shutil.rmtree(temp_dir)
            except Exception:
                for filename in os.listdir(temp_dir):
                    file_path = os.path.join(temp_dir, filename)
                    try:
                        if os.path.isfile(file_path) or os.path.islink(file_path):
                            os.unlink(file_path)
                        elif os.path.isdir(file_path):
                            shutil.rmtree(file_path)
                    except Exception:
                        pass
        os.makedirs(temp_dir, exist_ok=True)

        total_count = len(selected_items)
        for idx, item in enumerate(selected_items):
            c_name = item["corp_name"]
            r_name = item["report_nm"]
            rcp_no = item["rcept_no"]

            if progress_callback:
                progress_callback(idx, total_count, f"⏳ {c_name} - {r_name} 다운로드 중...")

            safe_c_name = re.sub(r'[\\/*?:"<>|]', "_", c_name).strip()
            safe_r_name = re.sub(r'[\\/*?:"<>|]', "_", r_name).strip()
            filename = f"{safe_c_name}_{safe_r_name}_{rcp_no}.zip"
            file_save_path = os.path.join(temp_dir, filename)

            download_original_document(
                rcept_no=rcp_no,
                save_path=file_save_path,
                api_key=self.api_key
            )

        if progress_callback:
            progress_callback(total_count, total_count, "📦 ZIP 파일 결합 압축 진행 중...")

        final_zip_path = os.path.join(base_dir, "original_documents_batch.zip")
        if os.path.exists(final_zip_path):
            try:
                os.remove(final_zip_path)
            except Exception:
                pass

        with zipfile.ZipFile(final_zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for root_dir, _, files in os.walk(temp_dir):
                for file in files:
                    file_path = os.path.join(root_dir, file)
                    zipf.write(file_path, os.path.relpath(file_path, temp_dir))

        try:
            shutil.rmtree(temp_dir)
        except Exception:
            pass

        return final_zip_path

    def run_batch_extract(
        self,
        corps: List[Dict[str, Any]],
        financial_statement_types: tuple,
        report_type: str,
        separate: bool,
        start_date: str,
        end_date: str,
        lang: str,
        save_dir: str,
        progress_callback=None
    ) -> List[Tuple[str, str, str]]:
        """
        여러 기업의 재무제표를 순차적으로 추출하여 엑셀 파일로 저장하고,
        각각의 실행 결과 로그(상태 메시지, 표시명, 저장경로)를 반환합니다.
        
        progress_callback: (current_idx, total_count, message) 형식으로 진행률을 전달하는 콜백
        """
        import pandas as pd
        os.makedirs(save_dir, exist_ok=True)
        results = []  # List[Tuple[message, display_name, file_path]]

        total = len(corps)
        for i, corp in enumerate(corps):
            code = corp["corp_code"]
            name = corp["corp_name"]

            if progress_callback:
                progress_callback(i, total, f"⏳ **{name}** ({code}) 추출 중...")

            try:
                fs = extract_financial_statement(
                    corp_code=code,
                    start_date=start_date,
                    end_date=end_date if end_date else None,
                    financial_statement_types=financial_statement_types,
                    separate=separate,
                    report_type=report_type,
                    lang=lang,
                    api_key=self.api_key,
                )

                safe_name = re.sub(r'[\\/*?:"<>|]', "_", name).strip()
                filename = f"{safe_name}_{code}.xlsx"
                filepath = os.path.join(save_dir, filename)

                if fs is not None and hasattr(fs, "_statements"):
                    exporter = FinancialStatementExporter(fs)
                    exporter.save(filepath)
                    msg = f"✅ **{name}** → `{filepath}`"
                    results.append((msg, f"{name} ({code})", filepath))
                elif fs is not None:
                    if isinstance(fs, pd.DataFrame):
                        fs.to_excel(filepath, index=False)
                        msg = f"✅ **{name}** → `{filepath}` (DataFrame)"
                        results.append((msg, f"{name} ({code})", filepath))
                    else:
                        msg = f"⚠️ **{name}** — 추출 완료, 저장 불가 (Exporter 미지원)"
                        results.append((msg, "", ""))
                else:
                    msg = f"⚠️ **{name}** — 추출 결과가 비어있습니다."
                    results.append((msg, "", ""))
            except Exception as err:
                msg = f"❌ **{name}** — 오류: {err}"
                results.append((msg, "", ""))

        if progress_callback:
            progress_callback(total, total, "완료!")

        return results


def main():
    parser = argparse.ArgumentParser(
        description='DART 기업 정보 및 재무제표 검색 도구 (Search DART corporation information and extract financial data using dart-fss.)'
    )

    # 기본 옵션 (Basic Options)
    parser.add_argument('--apikey', help='Open DART API 키 (또는 환경변수 DART_API_KEY 설정)')
    parser.add_argument('--update', action='store_true', help='최신 기업코드 목록 다운로드 및 캐싱')
    parser.add_argument('--cache', help='로컬 기업 목록 캐시 파일 경로')
    
    # 기업 검색 옵션 (Company Search Options)
    parser.add_argument('--keyword', help='기업명 또는 기업코드로 검색할 키워드')
    parser.add_argument('--market', default='YKNE', help='시장 필터: Y=코스피, K=코스닥, N=코넥스, E=기타 (기본값: YKNE)')
    parser.add_argument('--max', type=int, default=20, help='검색 결과 최대 개수 (기본값: 20)')
    parser.add_argument('--exact', action='store_true', help='기업명 정확 매칭 여부')
    
    # 재무보고서 옵션 (Financial Report Options)
    # [필수] 기업코드
    parser.add_argument('--corp-code', help='[필수] DART 기업코드 (8자리). 예) 00126380 (삼성전자)')
    # [필수] 검색 시작일
    parser.add_argument('--bgn-de', help='[필수] 검색 시작일자 (YYYYMMDD 형식). 예) 20200101')
    # [선택] 검색 종료일
    parser.add_argument('--end-de', help='[선택] 검색 종료일자 (YYYYMMDD 형식). 미지정시 오늘 기준')
    
    # 재무보고서 조회
    parser.add_argument('--list-reports', action='store_true', 
                       help='[--corp-code, --bgn-de 필수] 기간 내 이용 가능한 재무보고서 목록 조회')
    
    # 재무제표 추출
    parser.add_argument('--extract-fs', action='store_true',
                       help='[--corp-code, --bgn-de 필수] 재무제표 추출 (엑셀 저장 가능)')
    parser.add_argument('--fs-type', default='bs,is,cis,cf',
                       help='추출할 재무제표 유형 (쉼표 구분): bs=재무상태표, is=손익계산서, cis=포괄손익계산서, cf=현금흐름표 (기본값: bs,is,cis,cf)')
    parser.add_argument('--report-type', default='annual', choices=['annual', 'half', 'quarter'],
                       help='보고서 유형: annual=연간, half=반기, quarter=분기 (기본값: annual)')
    parser.add_argument('--separate', action='store_true',
                       help='개별재무제표 추출 (미지정시 연결재무제표)')
    parser.add_argument('--save-excel', help='추출된 재무제표를 엑셀 파일로 저장 (파일 경로)')
    parser.add_argument('--include-labels', action='store_true',
                       help='엑셀 저장 시 Labels_ (계정명 매핑) 시트도 함께 저장 (기본값: 수치 데이터 시트만 저장)')
    
    # 공시 원본문서 다운로드
    parser.add_argument('--download-doc', help='공시 원본문서 다운로드 (접수번호 지정). 예) 20240318000782')
    parser.add_argument('--save-path', help='다운로드한 원본문서(ZIP)를 저장할 경로 (파일 또는 폴더 경로)')

    args = parser.parse_args()

    try:
        set_api_key(args.apikey)
    except ValueError as exc:
        print(f'오류: {exc}', file=sys.stderr)
        parser.print_help()
        sys.exit(1)

    # [1] 기업 목록 업데이트
    if args.update:
        print('기업 목록을 다운로드 중입니다...')
        corp_list = download_corp_list(api_key=args.apikey, cache_file=args.cache, force=True)
        cache_path = build_cache_path(args.cache)
        print(f'완료: {len(corp_list)}개 기업 정보가 {cache_path} 에 저장되었습니다.')
        sys.exit(0)

    # [2] 기업 검색
    if args.keyword:
        keyword = args.keyword.strip()
        if not keyword:
            print('오류: 검색 키워드가 비어있습니다.', file=sys.stderr)
            sys.exit(1)

        if args.exact:
            # 정확한 기업명 매칭
            corp_list = load_cached_corp_list(cache_file=args.cache)
            if not corp_list:
                corp_list = download_corp_list(api_key=args.apikey, cache_file=args.cache, force=False)
            results = [corp for corp in corp_list if corp.get('corp_name') == keyword]
        else:
            results = search_corp_keyword(
                keyword,
                api_key=args.apikey,
                cache_file=args.cache,
                market=args.market,
                max_results=args.max,
                force_update=False,
            )

        if not results:
            print(f'검색 결과: 키워드 "{keyword}"에 해당하는 기업이 없습니다.')
            sys.exit(0)

        print(f'\n검색 결과: "{keyword}" 키워드로 {len(results)}개 기업을 찾았습니다.\n')
        for corp in results:
            print_corp_summary(corp)
        sys.exit(0)

    # [3] 재무보고서 목록 조회
    if args.list_reports:
        if not args.corp_code or not args.bgn_de:
            print('오류: --list-reports 사용 시 --corp-code 와 --bgn-de 필수입니다.', file=sys.stderr)
            print('예시: python DART_OpenAPI.py --corp-code 00126380 --bgn-de 20200101 --list-reports', file=sys.stderr)
            sys.exit(1)
        
        try:
            print(f'보고서를 조회 중입니다 (기업코드: {args.corp_code}, 기간: {args.bgn_de}~{args.end_de or "오늘"})...')
            reports = get_available_financial_reports(
                args.corp_code,
                args.bgn_de,
                args.end_de,
                api_key=args.apikey,
            )
            if not reports:
                print(f'조회 결과: 기업코드 {args.corp_code}에 해당하는 재무보고서가 없습니다.')
                sys.exit(0)
            
            print(f'\n조회 결과: {len(reports)}개의 재무보고서를 찾았습니다.\n')
            for idx, report in enumerate(reports, 1):
                print(f'[{idx}] {report.get("report_nm", "N/A")} ({report.get("reprt_code", "N/A")})')
                print(f'    기간: {report.get("bgn_de", "N/A")} ~ {report.get("end_de", "N/A")}')
                print(f'    접수번호: {report.get("rcept_no", "N/A")}')
                print()
            sys.exit(0)
        except Exception as e:
            print(f'오류: {e}', file=sys.stderr)
            sys.exit(1)

    # [4] 재무제표 추출
    if args.extract_fs:
        if not args.corp_code or not args.bgn_de:
            print('오류: --extract-fs 사용 시 --corp-code 와 --bgn-de 필수입니다.', file=sys.stderr)
            print('예시: python DART_OpenAPI.py --corp-code 00126380 --bgn-de 20200101 --extract-fs --save-excel output.xlsx', file=sys.stderr)
            sys.exit(1)
        
        try:
            fs_types = tuple(t.strip() for t in args.fs_type.split(','))
            
            print(f'재무제표를 추출 중입니다 (기업코드: {args.corp_code}, 기간: {args.bgn_de}~{args.end_de or "오늘"})...')
            print(f'제표 유형: {", ".join(fs_types)}, 보고서: {args.report_type}')
            
            fs = extract_financial_statement(
                args.corp_code,
                args.bgn_de,
                args.end_de,
                financial_statement_types=fs_types,
                separate=args.separate,
                report_type=args.report_type,
                api_key=args.apikey,
            )
            
            print(f'\n완료: 기업코드 {args.corp_code}의 재무제표를 성공적으로 추출했습니다.')
            
            if args.save_excel:
                if fs is not None and hasattr(fs, "_statements"):
                    exporter = FinancialStatementExporter(fs)
                    exporter.save(args.save_excel, include_labels=args.include_labels)
                    print(f'엑셀 저장: {args.save_excel}')
                else:
                    raise TypeError('추출된 재무제표가 유효하지 않거나 비어있습니다.')
            else:
                print('팁: --save-excel <파일경로> 옵션으로 엑셀 파일로 저장할 수 있습니다.')
            
            sys.exit(0)
        except Exception as e:
            print(f'오류: {e}', file=sys.stderr)
            sys.exit(1)

    # [5] 공시 원본문서 다운로드
    if args.download_doc:
        rcept_no = args.download_doc.strip()
        if not rcept_no:
            print('오류: --download-doc 사용 시 접수번호가 필요합니다.', file=sys.stderr)
            sys.exit(1)
            
        save_path = args.save_path or f"{rcept_no}.zip"
        
        try:
            print(f'공시 원본문서를 다운로드 중입니다 (접수번호: {rcept_no})...')
            saved_file = download_original_document(
                rcept_no=rcept_no,
                save_path=save_path,
                api_key=args.apikey,
            )
            print(f'완료: 원본문서가 성공적으로 다운로드되었습니다.')
            print(f'저장 경로: {saved_file}')
            sys.exit(0)
        except Exception as e:
            print(f'오류: {e}', file=sys.stderr)
            sys.exit(1)

    # 옵션이 없으면 도움말 표시
    print('오류: 실행할 작업을 지정해주세요.', file=sys.stderr)
    print('\n주요 사용법:', file=sys.stderr)
    print('  1. 기업 목록 업데이트:  python DART_OpenAPI.py --update', file=sys.stderr)
    print('  2. 기업 검색:          python DART_OpenAPI.py --keyword 삼성', file=sys.stderr)
    print('  3. 보고서 조회:        python DART_OpenAPI.py --corp-code 00126380 --bgn-de 20200101 --list-reports', file=sys.stderr)
    print('  4. 재무제표 추출:      python DART_OpenAPI.py --corp-code 00126380 --bgn-de 20200101 --extract-fs --save-excel out.xlsx', file=sys.stderr)
    print('  5. 원본문서 다운로드:  python DART_OpenAPI.py --download-doc 20240318000782 --save-path out_doc.zip', file=sys.stderr)
    print()
    parser.print_help()
    sys.exit(1)


if __name__ == '__main__':
    main()
